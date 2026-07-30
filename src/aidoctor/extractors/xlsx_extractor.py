"""XLSX extraction, one section per worksheet.

Spreadsheets are the format most often mangled by naive extraction. Two choices
matter here:

* **The header row is repeated onto every row.** A retrieved chunk saying
  ``EMEA | 1200`` is useless; ``Region: EMEA | Amount: 1200`` is answerable. The
  header is the only thing that makes a cell value mean anything, and chunking
  will eventually separate the two unless they are joined at extraction.
* **Formulas are read as cached values**, not as ``=SUM(B2:B9)``. openpyxl is
  opened with ``data_only=True``; the formula text is not the answer to a
  question about the number.

Empty rows and fully-empty columns are dropped, because a sparse sheet otherwise
produces chunks that are mostly separators.
"""

from __future__ import annotations

from pathlib import Path

from aidoctor.extractors.base import ExtractionError
from aidoctor.models.document import Document, Section, SourceType

MAX_ROWS_PER_SHEET = 2000


class XlsxExtractor:
    source_type = SourceType.XLSX
    extensions = (".xlsx", ".xlsm")

    def extract(self, path: Path, doc_id: str) -> Document:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency guard
            raise ExtractionError("openpyxl is required to read .xlsx files") from exc

        try:
            # data_only: we want 1200, not "=SUM(B2:B9)"
            workbook = load_workbook(str(path), data_only=True, read_only=True)
        except Exception as exc:
            raise ExtractionError(f"Could not open {path.name}: {exc}") from exc

        sections: list[Section] = []
        truncated: list[str] = []
        for sheet in workbook.worksheets:
            rows = [
                [("" if cell is None else str(cell).strip()) for cell in row]
                for row in sheet.iter_rows(values_only=True)
            ]
            rows = [r for r in rows if any(r)]
            if not rows:
                continue
            if len(rows) > MAX_ROWS_PER_SHEET + 1:
                truncated.append(sheet.title)
                rows = rows[: MAX_ROWS_PER_SHEET + 1]

            header, *body = rows
            lines: list[str] = []
            for row in body:
                pairs = [
                    f"{header[i] or f'col{i + 1}'}: {value}"
                    for i, value in enumerate(row)
                    if value and i < len(header)
                ]
                if pairs:
                    lines.append(" | ".join(pairs))
            if not lines:  # header-only sheet still carries meaning
                lines = [" | ".join(v for v in header if v)]
            sections.append(
                Section(text="\n".join(lines), label=f"sheet {sheet.title}", ordinal=len(sections))
            )
        workbook.close()

        if not sections:
            raise ExtractionError(f"{path.name} contained no populated cells")
        metadata = {"sheets": str(len(sections))}
        if truncated:
            metadata["truncated_sheets"] = ",".join(truncated)
        return Document(
            doc_id=doc_id,
            filename=path.name,
            source_type=self.source_type,
            sections=sections,
            metadata=metadata,
        )
